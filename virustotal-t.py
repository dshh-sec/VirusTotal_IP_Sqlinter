import os
import random
import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

# 创建一个参数对象，用来控制chrome以无界面模式打开
chrome_options = Options()
chrome_options.add_argument('--headless')  # 隐藏可视化界面
chrome_options.add_argument('--disable-gpu')  # 谷歌文档提档需要加上这个属性来规避bug
chrome_options.add_argument('blink-setting=imagesEnabled=false')  # 不加载图片
chrome_options.add_argument("disable-blink-features=AutomationControlled")  # 就是这一行告诉chrome去掉了webdriver痕迹
# chrome_options.add_argument("--proxy-server=socks://45.43.37.243:10802")  # 设置代理
chrome_options.add_argument("--proxy-server=http://127.0.0.1:10802")
chrome_options.add_argument('--log-level=3')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 不打印小于该级别的错误信息


def url_create(ip):
    check_ur = url_start + ip.strip("\n") + url_end
    return check_ur


def js_build():
    """构造js语句"""
    # js_conde = js + '.querySelector("div > vt-ui-expandable.resolutions > span > vt-ui-resolution-list").shadowRoot.querySelector("div > div.content > div:nth-child(4)")'
    pass


def init_excel(filename):
    """创建.xlsx表格，并初始化内容"""
    wb = openpyxl.Workbook()
    filename = str(filename.strip("\n"))
    filename = filename + ".xlsx"
    ws = wb.create_sheet(index=0, title="result")
    head = ['序号', '时间', '域名']
    for i in range(0, 3):
        ws.cell(1, i + 1).value = head[i]
        wb.save(filename)
    return str(filename)


def save_excel(data, filename):
    """保存数据"""
    wb_save = openpyxl.load_workbook(filename)
    ws_save = wb_save.active
    for rows in range(len(data)):
        ws_save.append(data[rows])
    # for rows in range(len(data)):
    #     ws_save.cell(rows + 1, 1).value = rows
    #     for lines in range(len(data[0])):
    #         ws_save.cell(rows + 2, lines + 2).value = str(data[rows][lines])
    wb_save.save(filename)


def execute(check_ur, web):
    """判断是否需要验证"""
    global js_conde_button

    web.get(check_ur)
    web.implicitly_wait(10)
    # time.sleep(random.random())
    print(web.title)
    # 判断人机身份验证
    if 'Captcha' in str(web.title):
        try:
            web.execute_script('return document.querySelector("#rc-anchor-container")').click()
        except Exception as ex:
            print("验证失败")
    else:
        click_element(web)


def click_element(web):
    # 获取“shadow_root”隐藏DOM树内容
    try:
        # shadow_button = web.execute_script(js_conde_button.format(random.randint(1, 3)))
        for i in range(5):
            shadow_button = web.execute_script(js_conde_button.format(random.randint(1, 3)))
            if type(shadow_button) is None:
                print("没有更多了")
            else:
                # shadow_button.click()
                web.execute_script("arguments[0].click();", shadow_button)
                web.implicitly_wait(10)
        get_element(web)
        #save_excel(data, filename=init_excel(ip))
    except Exception as err:
        print(err)
    # finally:
    #     dd = web.execute_script(
    #         'return document.querySelector("#view-container > ip-address-view").shadowRoot.querySelector("#relations").shadowRoot.querySelector("div > vt-ui-expandable.resolutions > span > div > vt-ui-button")')
        # print(dd)
        # print(dd.get_attribute("class"))
        # print(dd.is_displayed())


def get_element(web):
    """获取网页元素"""
    result = []
    count = 1
    while True:
        js_conde_element = js_conde_content.format(count)
        shadow_list = web.execute_script(js_conde_element)
        if shadow_list:
            content_time = shadow_list.find_element_by_class_name('regular-field')
            content_domain = shadow_list.find_element_by_class_name('long-field')
            content_list = [count, str(content_time.text), str(content_domain.text)]
            result.append(content_list)
            print(str(content_time.text) + '\t' + str(content_domain.text))
        else:
            break
        count += 1
    # print("result--->", result)
    save_excel(result, filename=init_excel(ip))


url_start = 'https://www.virustotal.com/gui/ip-address/'
url_end = '/relations'
file_path = os.getcwd() + '\\ip.txt'
js_conde_content = 'return document.querySelector("#view-container > ip-address-view").shadowRoot.querySelector("#relations").shadowRoot.querySelector("div > vt-ui-expandable.resolutions > span > vt-ui-resolution-list").shadowRoot.querySelector("div > div.content > div:nth-child({})")'
js_conde_button: str = 'return document.querySelector("#view-container > ip-address-view").shadowRoot.querySelector("#relations").shadowRoot.querySelector("div > vt-ui-expandable.resolutions > span > div > vt-ui-button").shadowRoot.querySelector("div > div.bounce{}")'

if __name__ == '__main__':
    if os.path.exists(file_path):
        # f = os.open(file_path, 'r')
        # for ip in open(file_path,'r'):
        with open(file_path, "r") as f:
            web = webdriver.Chrome(options=chrome_options,
                                   executable_path=r'C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe')
            for ip in f.readlines():
                execute(url_create(ip), web)
        web.close()
        web.quit()

    else:
        print("[-]:ip.txt 文件不存在！！！")
        time.sleep(random.random())
        exit()
